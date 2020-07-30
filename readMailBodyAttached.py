import win32com.client as win32
from outlook_msg import Message
import openpyxl


outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts

def lectura(fname,xlsx):
    cv_No_Processed_SPOT=[]
    newCont=''
    line=''
    flag=0
    linePointer=9
    with open(fname) as msg_file:
        msg = Message(msg_file)
    content=msg.body

    for c in content:
        if c!='\n' and c!='\r' and c!='\t':
            line+=c
        else:
            flag=1
            pass
        if flag==1:
            if c!='\n' and c!='\r' and c!='\t':
                lastChar=line[-1]
                line=line[:-1]
                line=line+'\n'
                line=line+lastChar
                flag=0
            else:
                pass
    line=line.split('\n')
    line = list(map(lambda x:x.upper(),line))
    print(line)
    data=[]
    fare=[]
    comData=[]
    for i in range(len(line)):
        if '$' in line[i]:
            fare.append(i)
    lastFareIndex=fare[-1]
    if 'NÚM SP ID OTM' in line:
        start=line.index('NÚM SP ID OTM')+10
    for i in range(start,lastFareIndex+1,10):
        singleData=[line[i],line[i+1],line[i+2],line[i+3],line[i+4],line[i+5],line[i+6],line[i+7],line[i+8],line[i+9],linePointer]
        print(singleData)
        linePointer+=1
        if ' ' in singleData:
            cv_No_Processed_SPOT.append(singleData)
        else:
            comData.append(singleData)
    comData.append(cv_No_Processed_SPOT)
    cv_No_Pro=comData[-1]
    comData=comData[:-1]
    inconsistantData=[]
    for i in comData:
        fareInd=i[-2].replace(' ','')
        i.pop(-2)
        i.insert(-2,fareInd)
        print(fareInd)
    wb = openpyxl.load_workbook(xlsx,read_only=True, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    print(comData,'aver')
    for i in range(len(comData)):
        col=3
        inconsistency=0
        row=comData[i][-1]
        print(row)
        for s in range(0,10):
            if s==8:
                fareXlsx=str(ws.cell(row,col).value)
                fareXlsx.replace(' ','')
                if fareXlsx!=comData[i][s]:
                    inconsistency=1
            elif str(ws.cell(row,col).value)!=comData[i][s]:
                inconsistency=1
                print(str(ws.cell(row,col).value)+','+comData[i][s])
            col+=1
        if inconsistency==0:
            pass
        # else:
        #     inconsistantData.append(comData[i][6])
        #     comData.pop(i)
    print(comData)
    print(inconsistantData)

def mail():
    inbox = outlook.GetDefaultFolder(6)                         # "6" refers to the index of a folder - in this case the inbox.                                      
    messages = inbox.Items
    for message in messages:
        if message.Unread==True:
            print(message.Subject)
            numberOfAttachments=len(message.Attachments)
            print(numberOfAttachments)
            if numberOfAttachments==2:
                file_name=r'C:\Users\aperalda\Downloads'+'\\'+message.Attachments[0].FileName
                file_name1=r'C:\Users\aperalda\Downloads'+'\\'+message.Attachments[1].FileName
                print(file_name)
                message.Attachments[0].SaveAsFile(file_name)
                message.Attachments[1].SaveAsFile(file_name1)
                docs=[file_name,file_name1]
                for i in docs:
                    if '.msg' in i:
                        msg=i
                    else:
                        xlsx=i
                print(msg)
                print('guardado')
                break

    lectura(msg,r'C:\Users\aperalda\Downloads\Formato de Solicitud de tarifa Individual 180620.xlsx')
    return message

def fw(correo):
    forward=correo.Forward()
    forward.SendUsingAccount=account[1]
    print(forward.SendUsingAccount)
    sender=correo.Sender
    print(sender)
    senders=str(sender)
    forward.To=senders
    forward.Send()
    print('ya')


mail()