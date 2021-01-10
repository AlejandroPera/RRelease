import asyncio, datetime, re, os, shutil,time, random, win32com.client as win32
from pyppeteer import launch
from datetime import timedelta
from openpyxl import load_workbook
from outlook_msg import Message

outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts

RATE_RECORD_MOD = "https://dsctmststr2.dhl.com/GC3/glog.webserver.finder.FinderServlet?ct=MTgxMzA2MDUwOTgwMTEwNTA2NQ%3D%3D&query_name=glog.server.query.rate.RateGeoQuery&finder_set_gid=MXPLAN.MX%20RECORD"
RATE_OFFERING_MOD = "https://dsctmststr2.dhl.com/GC3/glog.webserver.finder.FinderServlet?ct=NzExMTkyNjc2NjMzNTM4ODk3OQ%3D%3D&query_name=glog.server.query.rate.RateOfferingQuery&finder_set_gid=MXPLAN.MX%20OFFERING"
SAMPLE_RID = "115F_6049735_THN"
globalSender = ""

#Cambiar variables segun la prueba
testEmail = "jesus.vasquezs@dhl.com"#"rodrigo.narvaez@dhl.com"
resourcesPath = r'C:\\Users\\jesushev\\Documents\\GitHub\\RateRelease\\'#r"C:\\Users\\josenarv\\source\\repos\\RateRelease\\" 
documentsPath =  r'D:\Descargas'#r"C:\\Users\\josenarv\\Documents\\"
visualDelay = 5000


#Function for reading restricted local credentials
def readLoginCred(filename):
    f = open(filename, "r")
    user = f.readline().strip()
    psswd = f.readline()
    output = [user,psswd]
    return output

async def copyRateRecord(page, browser, new_rateOfferingID, origin, serviceProvider,unityType,provinceCode, population,fareType,fare):
    await page.goto(RATE_RECORD_MOD)
    record_field = await page.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type=text]:nth-child(3)")
    print("COLOCANDO ELEMENTOS PUNTO 5")
    await record_field.type(origin+"%"+unityType+"%"+provinceCode+"%"+population) #Aqui
    
    isActive_drop = await page.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(2) > td > div > select > option:nth-child(2)")
    search_button = await page.J("#search_button")
    await isActive_drop.click()
    await search_button.click()

    #3
    await page.waitForNavigation()
    await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    #Extracts innerText of element
    total_results = int (await page.evaluate('(element) => element.innerText', element))
    if total_results > 0:
        print("RATE RECORD: HAY DATOS")
        cb = await page.J("#rgSGSec\\.2\\.1\\.1\\.1 > input")
        copy_btn = await page.J("#resultsPage\\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(3) > table > tbody > tr > td:nth-child(6) > div > a")
        await cb.click()
        sinFrame = browser.targets()[len(browser.targets()) - 1]
        await page.waitFor(3000)
        await copy_btn.click()

        i=0
        print(i,browser.targets())
        conFrame = browser.targets()[len(browser.targets()) - 1]
        while (sinFrame == conFrame):
            
            await page.waitFor(500)
            conFrame = browser.targets()[len(browser.targets()) - 1]
            print(i,browser.targets())
            i+=1

        popUp_page = await conFrame.page()

        # A PARTIR DE AQUI

        print("Total frames:", len(popUp_page.frames))
        await popUp_page.waitFor("html > frameset > frame:nth-child(2)")
        frames = popUp_page.frames
        popUp = popUp_page.frames[2]

        #added 3 lines until 68
        await popUp.waitFor(5000)
        offeringID = await popUp.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type='text']:nth-child(3)")
        template_yes = await popUp.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(7) > td > div > table > tbody > tr > td:nth-child(1) > div > input[type='radio']")
        search_button = await popUp.J("#search_button")
        
        await offeringID.type(new_rateOfferingID)
        await search_button.click()

        await popUp.waitFor(visualDelay)
        
        await popUp.waitFor("#finished_button")
        finish_button = await popUp.J("#finished_button")
        await finish_button.click()
        
        await popUp.waitFor(visualDelay)

        await popUp.waitFor("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(1) > div > input[type='text']:nth-child(3)")
        rateRecordID = await popUp.J("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(1) > div > input[type='text']:nth-child(3)") 
        sourceRegion = await popUp.J("#rate_geo\\/x_lane\\/source\\/region_xid")
        destinationRegion = await popUp.J("#rate_geo\\/x_lane\\/destination\\/region_xid")
        active = await popUp.J("#rate_geo\\/is_active")
        attr_button = await popUp.J('#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a')
        
        #COLOCAR IF PARA TARIFA SPOT
        await rateRecordID.type(new_rateOfferingID+"-"+"CON_"+provinceCode+"_"+population)
        #await sourceRegion.type(origin)
        #await destinationRegion.type("CON_"+provinceCode+"_"+population)
        
        #VERIFICAR SI ESTA ACTIVA
        await active.click()
        
        await attr_button.click()
        await popUp.waitFor('#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a')
        rate_costs_button = await popUp.J('#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a')
        await rate_costs_button.click()
        await popUp.waitFor('#table_sec6_grid > tbody > tr > td:nth-child(4) > table > tbody > tr:nth-child(2) > td') #TIMEOUT-------------------------------------------------
        charge = await popUp.J('#table_sec6_grid > tbody > tr > td:nth-child(4) > table > tbody > tr:nth-child(2) > td')
        string = await popUp.evaluate('(charge) => charge.innerText', charge)
        clean_charge = string.split('.')[0].split('Charge ')[0]
        print('COSTO EN SISTEMA DECLARADO:',clean_charge)
        
        print("EXITOSO COPY RATE OFF:",rateID)
        
        await page.waitFor(700000)
    else:
        print("): RATE RECORD: SIN DATOS ")  
    return 1

async def rateOffPop(page, popUp,  origin,serviceProvider,unityType):
    print("EN POPUP RATE OFFERING")
    await popUp.waitFor('#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(1) > div > input[type="text"]:nth-child(4)')
    
    offeringID = await popUp.J('#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(1) > div > input[type="text"]:nth-child(4)')
    serviceProviderID = await popUp.J('#rate_offering\\/service_provider\\/xid')
    version = await popUp.J('#rate_offering\\/rate_version\\/xid')
    active = await popUp.J('#rate_offering\\/is_active')
    transportMode = await popUp.J('#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(2) > td:nth-child(3) > div > div:nth-child(3) > select')
    attributes_btn = await popUp.J('#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a')
    
    new_rate_offering_id =origin+"L_"+serviceProvider+"_"+unityType+str(random.randint(0,1000))
    await offeringID.type(new_rate_offering_id)
    #Espacio para verificar ESCENARIO 2 PASO 3
    #
    #
    await attributes_btn.click()
    
    await popUp.waitFor('body > div.bodyHeaderCont > table > tbody > tr > td:nth-child(2) > table > tbody > tr > td:nth-child(3) > div > a')
    finished_btn = await popUp.J('body > div.bodyHeaderCont > table > tbody > tr > td:nth-child(2) > table > tbody > tr > td:nth-child(3) > div > a')
    
    await finished_btn.click()

    await popUp.waitFor(2000)
    if (True): #Check if success pending
        success_id = await popUp.J('#successData > table:nth-child(6) > tbody > tr > td:nth-child(2) > span')
        rateID = await popUp.evaluate('(success_id) => success_id.innerText', success_id)
        print("EXITOSO COPY RATE OFF:",rateID)
        await popUp.waitFor(5000)
        await page.close()
        return rateID
    else:
        return 0

async def copyRateOffering(page, browser,origin, serviceProvider,unityType,provinceCode, population,fareType,fare):
    await page.goto(RATE_OFFERING_MOD)
    r_offering_id = await page.J('#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type="text"]:nth-child(3)')
    yes = await page.J('#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(4) > td > div > table > tbody > tr > td:nth-child(1) > div > input[type="radio"]')
    search_button = await page.J('#search_button')
    await r_offering_id.type(origin+"%"+unityType+"%")
    await yes.click()
    await search_button.click()
    await page.waitForNavigation()

    #3 Validate every rate via Rate Offering ID
    await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    #Extracts innerText of element
    total_results = int (await page.evaluate('(element) => element.innerText', element))
    mouse = page.mouse
    if total_results > 0:
        print("COPY RATE OFF: EXITOSA BUSQUEDA")
        box = await page.J('#rgSGSec\\.2\\.1\\.1\\.1 > input')
        actions =  await page.J('#resultsPage\\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(3) > table > tbody > tr > td:nth-child(5) > div > a')
        await box.click()
        await actions.click()
        await page.waitFor(1000)
        
        sinFrame = browser.targets()[len(browser.targets()) - 1]
        await mouse.click(355,149)
        await page.waitFor(3000)
        conFrame = browser.targets()[len(browser.targets()) - 1]
        while (sinFrame == conFrame):
            await page.waitFor(500)
            conFrame = browser.targets()[len(browser.targets()) - 1]

        popUp_page = await conFrame.page()
        popUp = popUp_page.frames[2]
        rateOfferingID = await rateOffPop(popUp_page,popUp,origin,serviceProvider,unityType)
        print("VOY")
        new_rateOfferingID=rateOfferingID.split('.')[1]
        await copyRateRecord(page,browser, new_rateOfferingID,origin, serviceProvider,unityType,provinceCode, population,fareType,fare)
        
    else:
        print("COPY RATE OFF: SIN RESULTADOS F")
        return 0

async def rateOffering(page, browser,origin, serviceProvider,unityType,provinceCode, population,fareType,fare):
    await page.goto(RATE_OFFERING_MOD)
    r_offering_id = await page.J('#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type="text"]:nth-child(3)')
    yes = await page.J('#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(4) > td > div > table > tbody > tr > td:nth-child(1) > div > input[type="radio"]')
    search_button = await page.J('#search_button')
    await r_offering_id.type(origin+"%"+serviceProvider+"%"+unityType+"XPO%")
    await yes.click()
    await search_button.click()
    await page.waitForNavigation()

    #3 Validate every rate via Rate Offering ID
    await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    total_results = int (await page.evaluate('(element) => element.innerText', element))
    if total_results > 0:
        print("RATEOFF: ENCONTRADOS FIN :(")
        return 1
    else:
        print("RATEOFF: SIN RESULTADOS")
        await copyRateOffering(page, browser,origin, serviceProvider,unityType,provinceCode, population,fareType,fare)  
        return 0

#-----------------------------------------------------------------------------------------------------
#Method to send Email
def sendEmail(address,body,subject):
    #mail.Sender
    o = win32.Dispatch("Outlook.Application")
    oacctouse = None
    for oacc in o.Session.Accounts:
        if oacc.SmtpAddress == "rodrigo.narvaez@dhl.com":
            oacctouse = oacc
            break
    Msg = o.CreateItem(0)
    if oacctouse:
        Msg._oleobj_.Invoke(*(64209, 0, 8, 0, oacctouse))  # Msg.SendUsingAccount = oacctouse

    Msg.To = address
    Msg.HTMLBody = body
    Msg.Subject= subject
    images_path = r'D:\Trabajo\NewRR\RateRelease\'
    Msg.Attachments.Add(Source=images_path+"DHL.png")    
    Msg.Send()

async def rateRecord(page, browser):
    await page.goto("https://dsctmststr2.dhl.com/GC3/glog.webserver.servlet.umt.Login")
    username_field = await page.J("body > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td > form > table > tbody > tr:nth-child(2) > td:nth-child(2) > input[type=text]")
    password_field = await page.J("body > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td > form > table > tbody > tr:nth-child(3) > td:nth-child(2) > input[type=password]")
    credentials = readLoginCred("cred.txt")
    await username_field.type(credentials[0])
    await password_field.type(credentials[1])
    await page.waitForNavigation()

    #2
    await page.goto(RATE_RECORD_MOD)
    record_field = await page.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type=text]:nth-child(3)")
    await record_field.type(SAMPLE_RID+"JJ")
    isActive_drop = await page.J("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(2) > td > div > select > option:nth-child(2)")
    search_button = await page.J("#search_button")
    await isActive_drop.click()
    await search_button.click()

    #3
    await page.waitForNavigation()
    await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
    #Extracts innerText of element
    total_results = int (await page.evaluate('(element) => element.innerText', element))
    if total_results > 0:
        cb = await page.J("#rgSGSec\.1\.1\.1\.1\.check")
        await cb.click()
    else:
        await rateOffering(page, browser)        
    return 1

#-----------------------------------------------------------------------------------------------------
#Module that manages emails
flagTime=0
formato=[]
tfares= "TarifMaster.xlsx" 

def txt_to_str(route):
    f = open(route, mode="r", encoding="utf-8")
    content = f.read()
    f.close()
    return str(content)

def createReply(email,mailBody,flag,acceptedCV):
    reply=email.Forward()
    sender='jesus.vasquezs@dhl.com'#'rodrigo.narvaez@dhl.com'#+'; alta.tarifas.sourcing@dhl.com'
    
    if flag==0:
        reply.HTMLBody = mailBody +'\n' +reply.HTMLBody
        sender='rpa.transport_@dhl.com'

    elif flag==1:
        init='<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="DHL.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV procesados</h2><table style="width:33%"><tr><th>CV Aceptados</th></tr>'
        for b in acceptedCV:
            init=init+'<tr><td>%s</td></tr>' %(b)
        init=init+'</table></body></html>'
        reply.HTMLBody = init + reply.HTMLBody

    elif flag==2:
        init='<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="DHL.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV sin tarifa en tarifario</h2><table style="width:33%"><tr><th>CV</th><th>Tipo de Unidad</th><th>Destino</th><th>Site</th></tr>'
        for f in acceptedCV:
            init=init+'<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>' %(f[6],f[3],f[5],f[2])
        init=init+'</table></body></html>'
        reply.HTMLBody = init + reply.HTMLBody
        sender='rpa.transport_@dhl.com'
    images_path = r"D:\Trabajo\NewRR\RateRelease\"
    reply.Attachments.Add(Source=images_path+"DHL.png")   
    reply.To=sender
    reply.Send()

    if flag==0:
        print(mailBody,' enviado')
    elif flag==1:
        print('CV aceptados enviado')
    elif flag==2:
        print('Sin tarifas en tarifario enviado')

def noProcessed(noProcessedCV,subject):

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    destinatarios='jesus.vasquezs@dhl.com'#'rodrigo.narvaez@dhl.com'#+'; alta.tarifas.sourcing@dhl.com'
    mail.To = destinatarios
    mail.Subject=subject
    f = open(r'C:\Users\jesushev\Documents\GitHub\RateRelease\mail.txt', 'w', encoding='UTF-8')
    f.write('<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"> <img src="DHL.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV no procesados</h2> <p style="font-family:verdana;">Esta es una alerta informativa sobre CV no capturados correctamente:</p><table style="width:100%"><tr><th>Inconsistencias</th><th>Falta de datos en correo adjunto</th><th>Falta de datos en csv</th></tr>')

    if len(noProcessedCV[1])>=len(noProcessedCV[2]) and len(noProcessedCV[1])>=len(noProcessedCV[3]):
        larger=len(noProcessedCV[1])
        while len(noProcessedCV[1])>len(noProcessedCV[2]):
            noProcessedCV[2].append('')
        while len(noProcessedCV[2])>len(noProcessedCV[3]):
            noProcessedCV[3].append('')
    elif len(noProcessedCV[2])>=len(noProcessedCV[1]) and len(noProcessedCV[2])>=len(noProcessedCV[3]):
        larger=len(noProcessedCV[2])
        while len(noProcessedCV[2])>len(noProcessedCV[1]):
            noProcessedCV[1].append('')
        while len(noProcessedCV[2])>len(noProcessedCV[3]):
            noProcessedCV[3].append('')
    else:
        larger=len(noProcessedCV[3])
        while len(noProcessedCV[3])>len(noProcessedCV[1]):
            noProcessedCV[1].append('')
        while len(noProcessedCV[3])>len(noProcessedCV[2]):
            noProcessedCV[2].append('')
    for i in range(larger):
        f.write("<tr><td>%s</td><td>%s</td><td>%s</td></tr>" %(noProcessedCV[1][i],noProcessedCV[2][i],noProcessedCV[3][i]))
    f.write('</table></body></html>')
    f.close()

    body=txt_to_str("mail.txt")
    mail.HTMLBody = body
    images_path = r"D:\Trabajo\NewRR\RateRelease\"

    mail.Attachments.Add(Source=  "DHL.png")
    mail.Send()
    print('No procesados enviado\n')

def not_in_tarif(array): #VALIDAR 0 no client, 1 no destination
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    destinatarios='jesus.vasquezs@dhl.com'#'rodrigo.narvaez@dhl.com'#+'; alta.tarifas.sourcing@dhl.com'
    mail.To = destinatarios
    mail.Subject='Sin datos en tarifario para cliente/destino'
    init='<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="DHL.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">Tarifas no encontradas por datos inconsistentes</h2><table style="width:33%"><tr><th>CV</th><th>Cliente</th><th>Destino</th></tr>'
    # if len(array[0])>=len(array[1]):
    #     while len(array[0])>len(array[1]):
    #         array[1].append('')
    # else:
    #     while len(array[1])>len(array[0]):
    #         array[0].append('')
    #print(array[0],'array 0')
    for i in range(len(array[0])):
        init=init+"<tr><td>%s</td><td>%s</td><td>%s</td></tr>" %(array[0][i][1],array[0][i][0],'')
    for i in range(len(array[1])):
        init=init+"<tr><td>%s</td><td>%s</td><td>%s</td></tr>" %(array[1][i][1],'',array[1][i][0])
    init=init+'</table></body></html>'
    mail.HTMLBody = init
    images_path = r"D:\Trabajo\NewRR\RateRelease\"
    mail.Attachments.Add(Source="DHL.png")
    mail.Send()       


def spotTipodeViaje(spot): #Encuentra el tipo de viaje par tarifas spot. Añade al último elemento de cada arreglo de tarifa el tipo de viaje.
    trf=load_workbook(tfares)
    ws = trf.worksheets[0]
    origin_Nomenclature_row = list(ws.rows)[1]
    destination_site_col=list(ws.columns)[2]
    unity_row=list(ws.rows)[2]
    dest=[cell.value for cell in destination_site_col]
    unities=[cell.value for cell in unity_row]
    spots=[]
    no_client=[]
    for f in spot:
        site=f[2].split()[0]
        state=f[5].split('/')[0].strip()
        destination=f[5].split('/')[1].strip()
        unity_type=f[3]
        destination_index=dest.index(destination)+1
        if site=='015':
            t_type=ws.cell(destination_index,4).value
        elif site=='009':
            t_type=ws.cell(destination_index,12).value
        elif site=='037':
            t_type=ws.cell(destination_index,27).value
        elif site=='140':
            t_type=ws.cell(destination_index,35).value
        elif site=='130':
            t_type=ws.cell(destination_index,43).value
        elif site=='139':
            t_type=ws.cell(destination_index,51).value
        elif site=='151':
            t_type=ws.cell(destination_index,59).value
        elif site=='187':
            t_type=ws.cell(destination_index,67).value
        elif site=='004':
            t_type=ws.cell(destination_index,75).value
        elif site=='051':
            t_type=ws.cell(destination_index,90).value
        elif site=='100':
            t_type=ws.cell(destination_index,105).value
        elif site=='108':
            t_type=ws.cell(destination_index,113).value
        elif site=='116':
            t_type=ws.cell(destination_index,121).value
        elif site=='065':
            t_type=ws.cell(destination_index,137).value
        elif site=='016':
            t_type=ws.cell(destination_index,145).value
        elif site=='083':
            t_type=ws.cell(destination_index,154).value
        elif site=='186':
            t_type=ws.cell(destination_index,161).value
        elif site=='002':
            t_type=ws.cell(destination_index,169).value
        elif site=='014':
            t_type=ws.cell(destination_index,178).value
        elif site=='019':
            t_type=ws.cell(destination_index,186).value
        elif site=='035':
            t_type=ws.cell(destination_index,195).value
        elif site=='024':
            t_type=ws.cell(destination_index,204).value
        elif site=='146':
            t_type=ws.cell(destination_index,213).value
        elif site=='027':
            t_type=ws.cell(destination_index,221).value
        elif site=='031':
            t_type=ws.cell(destination_index,229).value
        elif site=='132':
            t_type=ws.cell(destination_index,238).value
        elif site=='115':
            t_type=ws.cell(destination_index,246).value
        elif site=='145':
            t_type=ws.cell(destination_index,254).value
        elif site=='182':
            t_type=ws.cell(destination_index,264).value
        elif site=='185':
            t_type=ws.cell(destination_index,272).value
        elif site=='130':
            t_type=ws.cell(destination_index,277).value
        elif site=='139':
            t_type=ws.cell(destination_index,285).value
        elif site=='128':
            t_type=ws.cell(destination_index,293).value
        elif site=='189':
            t_type=ws.cell(destination_index,301).value
        elif site=='124':
            t_type=ws.cell(destination_index,309).value
        elif site=='50':
            t_type=ws.cell(destination_index,317).value
        elif site=='148':
            t_type=ws.cell(destination_index,325).value
        else:
            no_client.append([site,f[6]])
            continue

        f.insert(len(f),t_type)
        spots.append(f)
    trf.close()
    not_in_tarif_array=[no_client,[]]
    if len(no_client)>0:
        not_in_tarif(not_in_tarif_array)
    return spots
            
def xlsx_to_Array(xlsx):

    wb = load_workbook(xlsx, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    row_count=len(ws['B'])
    #processed_and_noProcessed=[[row_count],[noProcessedCV]] ### here
    #return processed_and_noProcessed
    filas=[]
    noProcessedCV=[]
    for i in range(9,row_count+1):
        fila=[]
        for a in range(3,13):
            if ws.cell(i,a).value != None:
                if ws.cell(i,a).value !='':
                    if a==9:
                        cv=str(ws.cell(i,a).value)
                        while len(cv)<8:
                            cv='0'+cv
                        fila.append(cv)
                        continue
                    else:
                        fila.append(ws.cell(i,a).value)
                else:
                    fila.append('')
            else:
                fila.append('')
        if '' in fila:
            if fila[6]!='':
                noProcessedCV.append(fila[6])
        else:
            filas.append(fila)
    processed_and_noProcessed=[filas,noProcessedCV]
    return processed_and_noProcessed


def outlookItem(fname):
    cv_SPOT_No_Processed=[]
    newCont=''
    line=''
    flag=0
    with open(fname) as msg_file:
        msg = Message(msg_file)
    content=msg.body

    for c in content:
        if c!='\n' and c!='\r':
            line+=c
        else:
            flag=1
            pass
        if flag==1:
            if c!='\n' and c!='\r':
                lastChar=line[-1]
                line=line[:-1]
                line=line+'\n'
                line=line+lastChar
                flag=0
            else:
                pass
    line=line.split('\n')
    line = list(map(lambda x:x.upper(),line))
    for ele in range(len(line)):
        if '\t' in line[ele]:
            line.insert(ele,' ')
            line[ele+1]=line[ele+1].replace('\t','')
    data=[]
    fare=[]
    comData=[]
    fullData=[]
    for i in range(len(line)):
        if '$' in line[i]:
            fare.append(i)
    lastFareIndex=fare[-1]
    if 'NÚM SP ID OTM' in line:
        start=line.index('NÚM SP ID OTM')+10
    for i in range(start,lastFareIndex+1,10):
        singleData=[line[i],line[i+1],line[i+2],line[i+3],line[i+4],line[i+5],line[i+6],line[i+7],line[i+8],line[i+9]]
        if ' ' in singleData:
            cv_SPOT_No_Processed.append(line[i+6])
        else:
            comData.append(singleData)
    for i in comData:
        fareInd=i[-1].replace(' ','')
        fareInd=fareInd.replace('$','')
        fareInd=fareInd.replace(',','')
        fareInd=fareInd[0:-3]
        i.pop(-1)
        i.insert(len(i),fareInd)
    
    fullData.append(comData)
    comData.append(cv_SPOT_No_Processed)
    return comData
 
def tariff(filas,filename,msg,mail):
    #tarifas
    cvProcessed=[]
    trf=load_workbook(tfares)
    ws = trf.worksheets[0]
    origin_Nomenclature_row = list(ws.rows)[1]
    destination_site_col=list(ws.columns)[2]
    unity_row=list(ws.rows)[2]
    dest=[cell.value for cell in destination_site_col]
    unities=[cell.value for cell in unity_row]
    tarifas=[]
    spot_fares=[]
    spot_not_valid=[]
    cv_No_Pro=[]
    fullData=[]
    no_tariff=[]
    no_destination=[]
    no_client=[]
    count=0
 
    try:
        spots=outlookItem(msg)
        spot_flag=1
        cv_No_Pro=spots[-1]
        for s in range(len(cv_No_Pro)):
            cv_s=cv_No_Pro[s]
            while len(cv_s)<8:
                cv_s='0'+cv_s
            cv_No_Pro[s]=cv_s
        spots=spots[:-1]
        for i in spots:
            cv=i[6]
            while len(cv)<8:
                cv='0'+cv
            i[6]=cv
            
    except:
        spot_flag=0
        print('No hay archivo .msg adjunto parae validar tarifas SPOT\n')
              
    for f in filas:
        if f[7] == 'TARIFA NORMAL':
            special_case_flag=0
            sites_special_origin=0
            flag_special_tarif=0
            
            if str(f[0])=='1909526':
                    if f[3]=='15T' or f[3]=='35T':
                        fare_CV='2650'
                        f.pop(-1)
                        f.insert(len(f),fare_CV)
                        tarifas.append(f)
                        flag_special_tarif=1
                    else: 
                        pass

            elif str(f[0])=='21613':
                fare_CV='99999'
                f.pop(-1)
                f.insert(len(f),fare_CV)
                tarifas.append(f)
                flag_special_tarif=1

            if not re.match('050|148', f[2][0:3]):
                sites_special_origin=1
            
            
            site=f[2].split()[0]
            state=f[5].split('/')[0].strip()
            destination=f[5].split('/')[1].strip()
            unity_type=f[3]
            if destination in ['LA PAZ', 'BENITO JUAREZ','CALERA']:
                special_case_flag=1
            if special_case_flag==0:
                try:
                    destination_index=dest.index(destination)+1
                except:
                    no_destination.append([f[5],f[6]])
                    continue
            else:
                if destination=='LA PAZ':
                    if state=='BCS':
                        destination_index=102
                    else:
                        destination_index=23
                elif destination=='CALERA':
                    if state=='ZAC':
                        destination_index=502
                    else:
                        destination_index=514
                else:
                    if state=='QTR':
                        destination_index=119
                    else:
                        destination_index=133
            # print(destination_index,'d')
            if site=='015':
                t_type=ws.cell(destination_index,4).value
                indexUnity=unities.index(unity_type,4,11)+1
            elif site=='009':
                t_type=ws.cell(destination_index,12).value
                indexUnity=unities.index(unity_type,12,26)+1
            elif site=='037':
                t_type=ws.cell(destination_index,27).value
                indexUnity=unities.index(unity_type,27,34)+1
            elif site=='140':
                t_type=ws.cell(destination_index,35).value
                indexUnity=unities.index(unity_type,35,42)+1
            elif site=='130':
                t_type=ws.cell(destination_index,43).value
                indexUnity=unities.index(unity_type,43,50)+1
            elif site=='139':
                t_type=ws.cell(destination_index,51).value
                indexUnity=unities.index(unity_type,51,58)+1
            elif site=='151':
                t_type=ws.cell(destination_index,59).value
                indexUnity=unities.index(unity_type,59,66)+1
            elif site=='187':
                t_type=ws.cell(destination_index,67).value
                indexUnity=unities.index(unity_type,67,74)+1
            elif site=='004':
                t_type=ws.cell(destination_index,75).value
                indexUnity=unities.index(unity_type,75,89)+1
            elif site=='051':
                t_type=ws.cell(destination_index,90).value
                indexUnity=unities.index(unity_type,90,104)+1
            elif site=='100':
                t_type=ws.cell(destination_index,105).value
                indexUnity=unities.index(unity_type,105,112)+1
            elif site=='108':
                t_type=ws.cell(destination_index,113).value
                indexUnity=unities.index(unity_type,113,120)+1
            elif site=='116':
                t_type=ws.cell(destination_index,121).value
                indexUnity=unities.index(unity_type,121,136)+1
            elif site=='065':
                t_type=ws.cell(destination_index,137).value
                indexUnity=unities.index(unity_type,137,144)+1
            elif site=='016':
                t_type=ws.cell(destination_index,145).value
                indexUnity=unities.index(unity_type,145,153)+1
            elif site=='083':
                t_type=ws.cell(destination_index,154).value
                indexUnity=unities.index(unity_type,154,160)+1
            elif site=='186':
                t_type=ws.cell(destination_index,161).value
                indexUnity=unities.index(unity_type,161,168)+1
            elif site=='002':
                t_type=ws.cell(destination_index,169).value
                indexUnity=unities.index(unity_type,169,177)+1
            elif site=='014':
                t_type=ws.cell(destination_index,178).value
                indexUnity=unities.index(unity_type,178,185)+1
            elif site=='019':
                t_type=ws.cell(destination_index,186).value
                indexUnity=unities.index(unity_type,186,194)+1
            elif site=='035':
                t_type=ws.cell(destination_index,195).value
                indexUnity=unities.index(unity_type,195,203)+1
            elif site=='024':
                t_type=ws.cell(destination_index,204).value
                indexUnity=unities.index(unity_type,204,212)+1
            elif site=='146':
                t_type=ws.cell(destination_index,213).value
                indexUnity=unities.index(unity_type,213,220)+1
            elif site=='027':
                t_type=ws.cell(destination_index,221).value
                indexUnity=unities.index(unity_type,221,228)+1
            elif site=='031':
                t_type=ws.cell(destination_index,229).value
                indexUnity=unities.index(unity_type,229,237)+1
            elif site=='132':
                t_type=ws.cell(destination_index,238).value
                indexUnity=unities.index(unity_type,238,245)+1
            elif site=='115':
                t_type=ws.cell(destination_index,246).value
                indexUnity=unities.index(unity_type,246,253)+1
            elif site=='145':
                t_type=ws.cell(destination_index,254).value
                indexUnity=unities.index(unity_type,254,263)+1
            elif site=='182':
                t_type=ws.cell(destination_index,264).value
                indexUnity=unities.index(unity_type,264,271)+1
            elif site=='185':
                t_type=ws.cell(destination_index,272).value
                indexUnity=unities.index(unity_type,272,276)+1
            elif site=='130':
                t_type=ws.cell(destination_index,277).value
                indexUnity=unities.index(unity_type,277,284)+1
            elif site=='139':
                t_type=ws.cell(destination_index,285).value
                indexUnity=unities.index(unity_type,285,292)+1
            elif site=='128':
                t_type=ws.cell(destination_index,293).value
                indexUnity=unities.index(unity_type,293,300)+1
            elif site=='189':
                t_type=ws.cell(destination_index,301).value
                indexUnity=unities.index(unity_type,301,308)+1
            elif site=='124':
                t_type=ws.cell(destination_index,309).value
                indexUnity=unities.index(unity_type,309,316)+1
            
            ##agregar sites con expecion 
            else:
                no_client.append([site,f[6]])
                continue
                
            fare_CV=ws.cell(destination_index,indexUnity).value

            if site=='50':
                t_type=ws.cell(destination_index,317).value
            elif site=='148':
                t_type=ws.cell(destination_index,325).value

            if fare_CV==None or fare_CV=='' or fare_CV==' ':
                no_tariff.append(f)
                
            else: #arreglar lul  
                if flag_special_tarif==0 or sites_special_origin==0: 
                    f.pop(-1)
                    f.insert(len(f),fare_CV)
                f.insert(len(f),t_type)
                tarifas.append(f)


        else:
            if spot_flag==1:
                for i in spots:
                    if i[6]==f[6]:
                        if i[0]==str(f[0]) and i[1]==f[1] and i[2]==f[2] and i[3]==f[3] and i[4]==str(f[4]) and i[9]==str(f[9]):
                            print(f)
                            spot_fares.append(f)
                        else:
                            #pass
                            if str(f[6]) not in cv_No_Pro:
                                spot_not_valid.append(f[6])
                count=0
                for s in spot_fares:
                    if str(f[6]) not in s:
                        count+=1
                if count==len(spot_fares) and f[6] not in cv_No_Pro and f[6] not in spot_not_valid:
                    print(f)
                    spot_not_valid.append(f[6])
            else:
                if str(f[6]) not in cv_No_Pro: 
                    spot_not_valid.append(f[6])

    trf.close()
    
    # for g in tarifas:
    #     cvProcessed.append(g[6])
    # for h in spot_fares:
    #     cvProcessed.append(h[6])
    not_in_tarif_array=[no_client,no_destination]
    if len(no_client)>0 or len(no_destination)>0:
        not_in_tarif(not_in_tarif_array)

    spot_fares=spotTipodeViaje(spot_fares)
    fullData.append(tarifas)
    fullData.append(spot_fares)
    fullData.append(spot_not_valid)
    fullData.append(cv_No_Pro)
    #print('fullDara',fullData)
    #if len(cvProcessed)>0:
        #createReply(mail,'Tarifas aceptadas: ',1,cvProcessed)
    if len(no_tariff)>0:
        createReply(mail,'Sin tarifa en tarifario: ',2,no_tariff)
    return fullData

def checkFormat(excel,mail):
    wb = load_workbook(excel, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    for i in range(2,13):
        formato.append(ws.cell(8,i).value.upper())
    if re.match('NO.?',formato[0]) and re.match('.*?ID.*?OTM', formato[1]) and re.match('.*?L[IÍ]NEA', formato[2]) and re.match('.*?PREDIO', formato[3]) and re.match('.*?UNIDAD', formato[4]) and re.match('C[P]?([ÓO]DIGO POSTAL)?', formato[5]) and re.match('POBLACI[ÓO]N', formato[6]) and re.match('C[V]?(ONTROL VEH[ÍI]CULAR)?', formato[7]) and re.match('.*?TARIFA', formato[8]) and re.match('AUTORIZA', formato[9]) and re.match('.*?IMPORTE', formato[10]):
        pass
    else:
        createReply(mail,"Formato Incorrecto",0,'')
        wb.close()
        return 0
    format=ws.cell(1,3).value.split()[-1] #[1:len(format)]
    f = open(r"S:\TRANSPORTE\LPC\TEMP\Beto\format.txt", "r")
    txt_content=f.read().split()
    validFormats = txt_content[:len(txt_content)]
    if format not in validFormats:
        createReply(mail,"Formato Incorrecto",0,'')
        wb.close()
        return 0

def validationTarif(filename,mail,msg):

    #if checkFormat(filename,mail)==0:
        #return 0

    processed_and_noProcessed=xlsx_to_Array(filename)
    filas=processed_and_noProcessed[0]
    noProcessedCV=processed_and_noProcessed[-1]

    tarifs_CV=tariff(filas,filename,msg,mail)
    tarifs_CV.append(noProcessedCV)
    return tarifs_CV
    
def validationSPOT(xlsxFile,arr_To_Validate,mail):
    
    if checkFormat(xlsxFile,mail)==0:
        return 0
    inconsistantData=[]
    consistants=[]
    cv_No_Pro=arr_To_Validate[-1]
    arr_To_Validate=arr_To_Validate[:-1]

    processed_and_noProcessed=xlsx_to_Array(xlsxFile)
    filas=processed_and_noProcessed[0]
    noProcessedCV=processed_and_noProcessed[-1]

    for element in arr_To_Validate:
        cv=str(element[6])
        while len(cv)<8:
            cv='0'+cv
        element[6]=cv

    for f in filas:
        if 'SPOT' in f[7]:
            for i in arr_To_Validate:
                if i[6]==f[6]:
                    if i[0]==str(f[0]) and i[1]==f[1] and i[2]==f[2] and i[3]==f[3] and i[4]==str(f[4]) and i[9]==str(f[9]): 
                        consistants.append(f)
                    else:
                        if f[6] not in cv_No_Pro:
                            inconsistantData.append(f[6])
            
            count=0
            for s in consistants:
                if str(f[6]) not in s:
                    count+=1
            if count==len(consistants) and f[6] not in cv_No_Pro:
                inconsistantData.append(f[6])

    # for element in consistants:
    #     cv=str(element[6])
    #     while len(cv)<8:
    #         cv='0'+cv
    #     element[6]=cv
    
    # for ele in range(len(inconsistantData)):
    #     cv=str(inconsistantData[ele])
    #     while len(cv)<8:
    #         cv='0'+cv
    #     inconsistantData[ele]=cv

    for el in range(len(cv_No_Pro)):
        cv=str(cv_No_Pro[el])
        while len(cv)<8:
            cv='0'+cv
        cv_No_Pro[el]=cv

    if len(arr_To_Validate)==0:
        arr_To_Validate.append([])
    
    spots_processed=[]
    # spots_accepted=[]
    # for y in consistants:
    #     spots_accepted.append(y[6])
    consistants=spotTipodeViaje(consistants) #Encuentra el tipo de viaje de spots
    spots_processed.append(consistants)
    spots_processed.append(inconsistantData)
    spots_processed.append(cv_No_Pro)
    spots_processed.append(noProcessedCV)
    #if len(spots_accepted)>0:
        #createReply(mail,'Spots aceptados: ',1,spots_accepted)
    return spots_processed
 
def retrieval():
    flag=0
    inbox = outlook.GetDefaultFolder(6)                         # "6" refers to the index of a folder - in this case the inbox.                                      
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)
    print('leyendo bandeja \n')
    msg=''
    for message in messages:
        if message.Unread==True and 'TARIFAS RPA' in message.Subject.upper():
            numberOfAttachments=len(message.Attachments)
            if numberOfAttachments<=2:
                globalSender = message.Sender
                print('Guardando\n')
                flag=1
                
                if numberOfAttachments==2:
                    file_name1=documentsPath+'\\'+ message.Attachments[0].FileName
                    file_name2=documentsPath+'\\'+ message.Attachments[1].FileName
                    message.Attachments[0].SaveAsFile(file_name1)
                    message.Attachments[1].SaveAsFile(file_name2)
                    docs=[file_name1,file_name2]
                    for i in docs:
                        if '.msg' in i:
                            msg=i
                        else:
                            file_name1=i
                else:
                    file_name1=documentsPath+'\\'+ message.Attachments[0].FileName
                    message.Attachments[0].SaveAsFile(file_name1)

                message.Unread = False
                print('Mensaje leído\n')
                break
            else:
                message.Unread=False
                time.sleep(1)
                print('Error en attachment\n')
                createReply(message,'El número de archivos adjuntos debe ser menor a dos: El archivo .xlsx y el .msg (En caaso de querer validar tarifas SPOT)',0,'')
                flag=0
                break

        elif message.Unread==True and 'SPOT RPA' in message.Subject.upper():
            numberOfAttachments=len(message.Attachments)
            if numberOfAttachments==2:
                globalSender = message.Sender
                print('Guardando\n')
                flag=2
                file_name1=documentsPath+'\\ '+ message.Attachments[0].FileName
                file_name2=documentsPath+'\\' + message.Attachments[1].FileName
                message.Attachments[0].SaveAsFile(file_name1)
                message.Attachments[1].SaveAsFile(file_name2)
                docs=[file_name1,file_name2]
                for i in docs:
                    if '.msg' in i:
                        msg=i
                    else:
                        xlsxFile=i
                message.Unread = False
                print('Mensaje leído\n')
                break
            else:
                message.Unread=False
                print('Error en attachment\n')
                createReply(message,'El número de archivos adjuntos debe ser igual a dos: El archivo .xlsx y el .msg',0,'')
                flag=0
                break
            
 
    if flag==1:
        flag=0
        print('Leyendo tarifas')
        final_fares_Array=validationTarif(file_name1, message,msg)
        if final_fares_Array==0:
            return 0
        fixed_array=[final_fares_Array[0]+final_fares_Array[1],final_fares_Array[2],final_fares_Array[3],final_fares_Array[4]]
        if len(final_fares_Array[-1])>0 or len(final_fares_Array[-2])>0 or len(final_fares_Array[-3])>0:
            noProcessed(fixed_array,'CV no procesados')
        os.remove(file_name1)
        print('Tarifas array: ')
        print(fixed_array)
        print(fixed_array[0])
        return fixed_array
    elif flag==2:
        flag=0
        final_fares_Array=validationSPOT(xlsxFile,outlookItem(msg),message)
        if final_fares_Array==0:
            return 0
        if len(final_fares_Array[-1])>0 or len(final_fares_Array[-2])>0 or len(final_fares_Array[-3])>0:
            noProcessed(final_fares_Array,'CV no procesados')
        os.remove(xlsxFile)
        os.remove(msg)
        print('Spot array:')
        print(final_fares_Array)
        return final_fares_Array
 
    else:
        print('No se encontró alta de tarifas\n')
#-----------------------------------------------------------------------------------------------------
#Login to OTM
async def rateRecordOTM(browser, page, acceptedFares, fareUniverse):
    #print(fares)
    SAMPLE_RID = []
    successFares = []
    failedFares = []
    origin = []
    serviceProvider = []
    unityType = []
    provinceCode = []
    population = []
    fareType = []
    fare = [] # $ the cost
    spot = False

    for i in range(len(acceptedFares)):
        origin.append(str(acceptedFares[i][2].split(" |")[0]))
        serviceProvider.append(str(acceptedFares[i][0]))
        unityType.append(str(acceptedFares[i][3]))
        provinceCode.append(str(acceptedFares[i][5].split(" /")[0]))
        if(provinceCode[i]=="CMX"):
            provinceCode[i]="MDF"
        population.append(str(acceptedFares[i][5].split("/ ")[1]))    
        fareType.append(str(acceptedFares[i][7]))
        fare.append(str(acceptedFares[i][9]))

        SAMPLE_RID.append(str(origin[i])+"%"+str(serviceProvider[i])+"%"+str(unityType[i])+"%"+str(provinceCode[i])+"%"+str(population[i]))
    #TEMPORAL DELETE OUTSIDE TESTING
    #origin=["009","115","115"]
    #serviceProvider= ["3095637","6049735","38877"]
    #unityType=["15T","THN","THN"]
    #provinceCode= ["MIC","SON","TMS"]
    #population=["TANGANCICUARO","CAJEME","EL_MANTE"]
    #fareType = ["TARIFA NORMAL","TARIFA NORMAL","TARIFA NORMAL"]
    #fare = ["3500","4300","3200"]       
    #SAMPLE_RID=["009%3095637%15T%MIC%TANGANCICUARO","115%6049735%THN%SON%CAJEME","115%38877%THN%TMS%EL_MANTE"]   
    
    for i in range(len(SAMPLE_RID)):
        await page.goto("https://dsctmststr2.dhl.com/GC3/glog.webserver.finder.FinderServlet?ct=MTgxMzA2MDUwOTgwMTEwNTA2NQ%3D%3D&query_name=glog.server.query.rate.RateGeoQuery&finder_set_gid=MXPLAN.MX%20RECORD")  
        print("PROBANDO CON RATE:", SAMPLE_RID[i])
        #await page.waitFor(3500)
        await page.type("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(1) > td > div > input[type=text]:nth-child(3)",SAMPLE_RID[i]) 
        await page.click("#bodyDataDiv > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(2) > td > div > select > option:nth-child(2)")
        await page.click("#search_button")
        await page.waitFor(2000)

        await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
        element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
        total_results = int (await page.evaluate('(element) => element.innerText', element))
        if total_results == 0:
            print("0")
            await rateOffering(page,browser,origin[i],serviceProvider[i],unityType[i],provinceCode[i], population[i],fareType[i],fare[i])
        else:
            print(" *** EXISTE, VALIDANDO CAMPOS ***")
            #3 Validate every rate via Rate Offering ID
            await page.waitFor("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
            element = await page.J("#resultsPage\:pgl2 > tbody > tr > td > div > table > tbody > tr > td:nth-child(2) > span:nth-child(2)")
            #Extracts innerText of element
            total_results = int (await page.evaluate('(element) => element.innerText', element))
            if total_results > 0:
                cb = await page.J("#rgSGSec\.1\.1\.1\.1\.check")
                await cb.click()
                element = await page.J("#rgPageSelectedTotal")
                #Total selected
                selected_total = int (await page.evaluate('(element) => element.innerText', element))
                #Click on the first one
                #for i in range(1,selected_total): #CHECK when it appears more than 1
                await page.click("#rgSGSec\\2e 2\\2e 2\\2e {}\\2e 1 > a".format(1))
                await page.waitFor(2000)

                #Control of pop up
                popup = await changingToPopUp(browser)
                await popup.waitFor(2200)
                
                frame=popup.frames[2] # Place on the frame containing the element
                
                #4 Validate Rate offering ID 
                element = await frame.J("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(1) > div > div.fieldData")
                offeringID = await frame.evaluate('(element) => element.textContent', element)
                splittedOffering = offeringID.split("_")
                firstPart =  list(splittedOffering[0])
                transportOID = firstPart[len(firstPart)-1]
                
                element = await frame.J("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(2) > td:nth-child(1) > div.fieldData > a")
                version = await frame.evaluate('(element) => element.textContent', element)
            
                element = await frame.J("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(2) > td:nth-child(4) > div.fieldData")
                equipmentGProfile = await frame.evaluate('(element) => element.textContent', element)

                element = await frame.J("#bodyDataDiv > div:nth-child(1) > table > tbody > tr:nth-child(1) > td:nth-child(6) > div.fieldData")
                transportMode = await frame.evaluate('(element) => element.textContent', element)

                element = await frame.waitForXPath('//*[@id="bodyDataDiv"]/div[1]/table/tbody/tr[2]/td[6]/div/div[2]/img/@alt')
                activeArrow = await frame.evaluate('(element) => element.textContent', element)

                transportM =""
                if transportOID=="F":
                    transportM="TLF"
                else:
                    transportM="TLL"

                if (activeArrow=="TRUE" and version=="ACTIVE" and (equipmentGProfile==unityType[i]) and (transportMode == transportM) ):
                    print("Rate Offering valid: {}".format(i))
                    await popup.close()

                    #Change again to the main page
                    popup = await changingToPopUp(browser)
                    await popup.waitFor(2200)

                    await popup.click("#rgSGSec\\2e 2\\2e 1\\2e {}\\2e 2 > a".format(1))
                    await popup.waitFor(2500)
                    #Control of new pop up
                    popup = await changingToPopUp(browser)
                    await popup.waitFor(2500)
                
                    frame=popup.frames[2] # Place on the frame containing the element

                    #rateRecordID
                    element = await frame.J("#readonly_rate_geo\\2f xid")
                    rateRecordID = await frame.evaluate('(element) => element.textContent', element)
                    
                    splittedRateRecord = rateRecordID.split("_")
                    firstPart =  (splittedRateRecord[0])
                    firstPartList= list(firstPart)

                    partNoRecord=''
                    partTripRecord=''
                    if(firstPartList[len(firstPartList)-1]=='F'):
                        partNoRecord = firstPart.partition('F')[0]
                        partTripRecord = firstPart.partition('F')[1]
                    elif(firstPartList[len(firstPartList)-1]=='L'):
                        partNoRecord = firstPart.partition('L')[0]
                        partTripRecord = firstPart.partition('L')[1]
                    
                    #rateOffering
                    element = await frame.J("#readonly_rate_geo\\2f rate_offering\\2f xid")
                    rateOfferingID = await frame.evaluate('(element) => element.textContent', element)
                    splittedRateOffering = rateRecordID.split("_")
                    firstPart =  (splittedRateOffering[0])
                    firstPartList= list(firstPart)

                    partNoOffering=''
                    partTripOffering=''
                    if(firstPartList[len(firstPartList)-1]=='F'):
                        partNoOffering = firstPart.partition('F')[0]
                        partTripOffering = firstPart.partition('F')[1]
                    elif(firstPartList[len(firstPartList)-1]=='L'):
                        partNoOffering = firstPart.partition('L')[0]
                        partTripOffering = firstPart.partition('L')[1]

                    element = await frame.J("#rate_geo\\2f x_lane\\2f source\\2f region_xid")
                    sourceRegion = (await frame.evaluate('(element) => element.value', element)).split(" ")[0]

                    element = await frame.J("#rate_geo\\2f x_lane\\2f destination\\2f region_xid")
                    destinationRegion = await frame.evaluate('(element) => element.value', element)   

                    fareDestinationRegion= "CON_"+provinceCode[i]+"_"+population[i]
                    
                    #The destinationRegion and sourceRegion must match with the fares from the mail
                    if(partNoOffering == partNoRecord and partTripOffering == partTripRecord and sourceRegion == origin[i] and destinationRegion == fareDestinationRegion):
                        print("Rate Record valid: {}".format(i))
                        #Rate Cost part
                        await frame.click("#rate\\2f RateGeoCost\\2e xsl > div.tabLabelCont.tabDim > span > a")
                        await frame.waitFor(2000)
                        if(fareType[i]=="TARIFA SPOT"):
                            spot = True

                        if(spot == True):                                
                            await frame.click("#table_sec6_grid > tbody > tr > td.gridBodyCell.gridBodyBtnsCell > table > tbody > tr > td:nth-child(1) > a > img")
                            await frame.waitFor(2000)

                            #Control of new pop up
                            popup = await changingToPopUp(browser)
                            await popup.waitFor(2000)

                            frame=popup.frames[2] # Place on the frame containing the element

                            amountOTM = await frame.J("#charge_amount_td > div > input.num")
                            #clear value before typing
                            await amountOTM.click()
                            await popup.keyboard.down("Control")
                            await popup.keyboard.press('KeyA')
                            
                            await popup.keyboard.up('Control')
                            await amountOTM.type(str(fare[i])) 
                            
                            #Change expiration date 
                            element = await frame.J("#rate_geo_cost\/expiration_date\:\:content ")  
                            expirationDate =  await frame.evaluate('(element) => element.value', element)    
                            expirationDate = datetime.strptime(expirationDate,'%Y-%m-%d')  
                            #print(expirationDate)             
                            now = datetime.now()
                            #print(now)
                            if (expirationDate >= now ):
                                print("Spot fare expired, reactivate it")                            
                                await expirationDate.click()
                                await popup.keyboard.down("Control")
                                await popup.keyboard.press('KeyA')
                            
                                await popup.keyboard.up('Control')
                                newDate = (now+timedelta(days=1)).strftime("%Y-%m-%d")
                                await expirationDate.type(str(now))
                            
                            #click on save
                            await frame.click("#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a")
                            await frame.waitFor(2000)

                            #Control of new pop up
                            popup = await changingToPopUp(browser)
                            await popup.waitFor(2000)

                            frame=popup.frames[2] # Place on the frame containing the element

                            #Finished spot fare
                            await frame.click("body > div.bodyHeaderCont > table > tbody > tr > td:nth-child(2) > table > tbody > tr > td:nth-child(3) > div > a")
                            print("All okey with the fare")
                            tempList = [acceptedFares[i][6],acceptedFares[i][0],"Tarifa Spot en orden"]
                            successFares.append(tempList)
                            await frame.waitFor(2200)
                            await popup.close()
                            #return to main page
                            popup = await changingToPopUp(browser)
                            await page.waitFor(2500)

                        else:                
                            element = await frame.J("#table_sec6_grid > tbody > tr > td:nth-child(4) > table > tbody > tr:nth-child(2) > td")
                            description = await frame.evaluate('(element) => element.textContent', element) #description with cost included
                            
                            splittedDescription = list(description.split(" "))                        
                            
                            #compare between OTM and the fares from the mail
                            if(splittedDescription[2]==fare[i]):
                                #Finished normal fare
                                tempList = [acceptedFares[i][6],acceptedFares[i][0],"Tarifa Normal en orden"]
                                successFares.append(tempList)
                                await frame.click("body > div.bodyHeaderCont > table > tbody > tr > td:nth-child(2) > table > tbody > tr > td:nth-child(3) > div > a")

                            else:
                                print("Your OTM fare doesn´t match the fare on the file")
                                #Add the correct fare
                                await frame.click("#table_sec6_grid > tbody > tr > td.gridBodyCell.gridBodyBtnsCell > table > tbody > tr > td:nth-child(1) > a > img")
                                await frame.waitFor(2700)

                                #Control of new pop up
                                popup = await changingToPopUp(browser)
                                await popup.waitFor(2900)

                                frame=popup.frames[2] # Place on the frame containing the element

                                amountOTM = await frame.J("#charge_amount_td > div > input.num")
                                #clear value before typing
                                await amountOTM.click()
                                await popup.keyboard.down("Control")
                                await popup.keyboard.press('KeyA')
                                
                                await popup.keyboard.up('Control')
                                await amountOTM.type(str(fare[i])) #suppose it's empty if it has somevalue add clear field part
                                #click on save
                                await frame.click("#bodyDataFooterContDiv > table > tbody > tr > td:nth-child(2) > div > a")
                                await frame.waitFor(2000)

                                #Control of new pop up
                                popup = await changingToPopUp(browser)
                                await popup.waitFor(2000)

                                frame=popup.frames[2] # Place on the frame containing the element

                                #Finished normal fare
                                await frame.click("body > div.bodyHeaderCont > table > tbody > tr > td:nth-child(2) > table > tbody > tr > td:nth-child(3) > div > a")
                                await frame.waitFor(2000)
                                print("All okey with the fare")
                                tempList = [acceptedFares[i][6].strip("0"),acceptedFares[i][0],"Tarifa Normal corregida"]
                                successFares.append(tempList)
                                await popup.close()
                                #return to main page
                                popup = await changingToPopUp(browser)
                                await page.waitFor(2000)                        
                                    
                    else:
                        print("Failed on Rate Record: {}".format(i))
                        tempList = [acceptedFares[i][6].strip("0"),acceptedFares[i][0],fareType[i],"Validación fallida en Rate Record"]
                        failedFares.append(tempList)
                        await popup.close()
                        #return to main page
                        popup = await changingToPopUp(browser)
                        await page.waitFor(2000)
                    
                else:
                    print("Failed on Rate Offering {}".format(i))
                    tempList = [acceptedFares[i][6].strip("0"),acceptedFares[i][0],fareType[i],"Validación fallida en Rate Offering"]
                    failedFares.append(tempList)
                    await popup.close()
                    #return to main page
                    popup = await changingToPopUp(browser)
                    await page.waitFor(2000)
            else:      
                print("RID not found on OTM System")
                tempList = [acceptedFares[i][6].strip("0"),acceptedFares[i][0],fareType[i],"RID no encontrado en OTM"]
                failedFares.append(tempList)
                await page.waitFor(1000)

    #Notifiy the applicant
    success='<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="DHL.png" alt="DHL LOGO.png"width="289", height="82"> <h2 style="font-family:verdana;text-align:center;">Tarifas Correctas</h2><table style="width:33%"><tr><th>Control Vehicular</th><th>Num SP ID OTM</th><th>Estatus</th></tr>'
    failed = '<style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="DHL.png" alt="DHL LOGO.png"width="289", height="82"> <h2 style="font-family:verdana;text-align:center;">Tarifas Incorrectas</h2><table style="width:33%"><tr><th>Control Vehicular</th><th>Num SP ID OTM</th><th>Estatus</th><th>Razón</th></tr>'
    body=''
    if len(successFares)>0:
        for i in range(len(successFares)):
            success += '<tr><td>{0}</td><td>{1}</td><td>{2}</td></tr>'.format(successFares[i][0],successFares[i][1],successFares[i][2])
        success +='</table>'
        body +=success
    if len(failedFares)>0:
        for i in range(len(failedFares)):
            failed += '<tr><td>{0}</td><td>{1}</td><td>{2}</td><td>{3}</td></tr>'.format(failedFares[i][0],failedFares[i][1],failedFares[i][2],failedFares[i][3])
        failed +='</table>'
        body +=failed
    body += '</body></html>'

    if len(successFares)>0 or len(failedFares)>0:                       
        sendEmail(testEmail,body,"Reporte Tarifas")     
    await browser.close()
      
async def puppet(fares):
    retries = 1 
    for i in range (retries):
        #try:
        browser = await launch(headless = False)
        page = await browser.newPage()

        #1 Login
        await page.goto("https://dsctmststr2.dhl.com/GC3/glog.webserver.servlet.umt.Login")
        credentials = readLoginCred("cred.txt")
        passw=await page.waitFor("[name='userpassword']")
        usernN=await page.waitFor("[name='username']")
        await usernN.type(credentials[0])
        await passw.type(credentials[1])        
        await page.click("[name='submitbutton']")
        await page.waitFor(1000)
        
        #2 Search rate and validate
        if fares!=0:
            await rateRecordOTM(browser, page, fares[0],fares) #mandar a llamar raterecord con solo las fares aceptadas.
        return 1    
        #except Exception as e:
        #    print("Error en puppet, reintentando", e)

async def testPuppet2():
    browser = await launch(headless = False)
    page = await browser.newPage()
    #1 Login
    await page.goto("https://dsctmststr2.dhl.com/GC3/glog.webserver.servlet.umt.Login")
    #2 Search rate and validate
    await rateRecord(page, browser) 
    return 1    
    '''try:
        browser = await launch(headless = False)
        page = await browser.newPage()

        #1 Login
        await page.goto("https://dsctmststr2.dhl.com/GC3/glog.webserver.servlet.umt.Login")
        credentials = readLoginCred("cred.txt")
        passw=await page.waitFor("[name='userpassword']")
        usernN=await page.waitFor("[name='username']")
        await usernN.type(credentials[0])
        await passw.type(credentials[1])        
        await page.click("[name='submitbutton']")
        await page.waitFor(1000)

        #2 Search rate and validate
        await rateRecord(page) 
        return 1    
    except Exception as e:
        print("Error en puppet, reintentando", e)
       ''' 
#-----------------------------------------------------------------------------------------------------
#Changing to pop up
async def changingToPopUp(browser):
    popLen = len(browser.targets())
    pop = browser.targets()[popLen-1]
    popup = await pop.page()
    return popup

#-----------------------------------------------------------------------------------------------------
#Scenario 1 Rate already exists
      
async def main():
    #Extract fare from email
    while 1:                    # ?
        fares = retrieval()
        if not fares is None:
            await puppet(fares) #Login on OTM and following steps                    


#asyncio.get_event_loop().run_until_complete(testPuppet2())
asyncio.get_event_loop().run_until_complete(main())