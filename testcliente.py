import win32com.client as win32
import datetime,re,os,shutil,time
from openpyxl import load_workbook
from outlook_msg import Message
 
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts
flagTime=0
formato=[]
tfares=r'D:\Descargas\TarifMaster.xlsx'
 
def outlookItem(fname):
    cv_SPOT_No_Processed=[]
    linePointer=9
    newCont=''
    line=''
    flag=0
    with open(fname) as msg_file:
        msg = Message(msg_file)
    content=msg.body

 
    for c in content:
        if c!='\n' and c!='\r':
            line+=c
        else:
            flag=1
            pass
        if flag==1:
            if c!='\n' and c!='\r':
                lastChar=line[-1]
                line=line[:-1]
                line=line+'\n'
                line=line+lastChar
                flag=0
            else:
                pass
    line=line.split('\n')
    line = list(map(lambda x:x.upper(),line))
    for ele in range(len(line)):
        if '\t' in line[ele]:
            line.insert(ele,' ')
            line[ele+1]=line[ele+1].replace('\t','')
    data=[]
    fare=[]
    comData=[]
    fullData=[]
    for i in range(len(line)):
        if '$' in line[i]:
            fare.append(i)
    lastFareIndex=fare[-1]
    if 'NÚM SP ID OTM' in line:
        start=line.index('NÚM SP ID OTM')+10
    for i in range(start,lastFareIndex+1,10):
        singleData=[line[i],line[i+1],line[i+2],line[i+3],line[i+4],line[i+5],line[i+6],line[i+7],line[i+8],line[i+9],linePointer]
        linePointer+=1
        if ' ' in singleData:
            cv_SPOT_No_Processed.append(line[i+6])
        else:
            comData.append(singleData)
    for i in comData:
        fareInd=i[-2].replace(' ','')
        fareInd=fareInd.replace('$','')
        fareInd=fareInd.replace(',','')
        fareInd=fareInd[0:-3]
        i.pop(-2)
        i.insert(-1,fareInd)
    
    fullData.append(comData)
    comData.append(cv_SPOT_No_Processed)
    return comData
 
def txt_to_str(route):
    f = open(route, mode="r", encoding="utf-8")
    content = f.read()
    f.close()
    return str(content)
 
def noProcessed(noProcessedCV):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    destinatarios=''
    mail.To = destinatarios
    mail.Subject='CV no procesados'
    #f = open(r'C:\Users\aperalda\Documents\adicionales\mail.txt', 'w', encoding='UTF-8')
    f.write('<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"> <img src="llpc.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV no procesados</h2> <p style="font-family:verdana;">Esta es una alerta informativa sobre CV no capturados correctamente:</p><table style="width:100%"><tr><th>CV/th></tr>')
    for i in range(len(noProcessedCV)):
        f.write("<tr> <td>%s</td> </tr>\n" %(noProcessedCV[i]))
    f.write('</table></body></html>')
    f.close()
    body=txt_to_str(r'C:\Users\aperalda\Documents\adicionales\mail.txt')
    mail.HTMLBody = body
    images_path = "C:\\Users\\aperalda\\Documents\\RAudit\\RateAudit\\img\\"
    mail.Attachments.Add(Source= images_path+"llpc.PNG")
    mail.Send()
    print('No procesados enviado\n')
 
def tariifario(filas,filename,msg):
 
    trf=load_workbook(tfares)
    ws = trf.worksheets[0]
    origin_Nomenclature_row = list(ws.rows)[1]
    destination_site_col=list(ws.columns)[2]
    unity_row=list(ws.rows)[2]
    dest=[cell.value for cell in destination_site_col]
    unities=[cell.value for cell in unity_row]
    tarifas=[]
    spot_fares=[]
    spot_not_valid=[]
    cv_No_Pro=[]
    fullData=[]
    count=0
 
    try:
        spots=outlookItem(msg)
        spot_flag=1
        cv_No_Pro=spots[-1]
        spots=spots[:-1]
    except:
        spot_flag=0
        print('No hay elementos para validar SPOT\n')
                
    for f in filas:
        if f[8] == 'TARIFA NORMAL':
            if not re.match('050|128|148', f[3][0:3]):
                special_case_flag=0
                site=f[3].split()[0]
                state=f[6].split('/')[0].strip()
                destination=f[6].split('/')[1].strip()
                unity_type=f[4]
                # print(unity_type,'Unity type')
                # print(site,'o')
                # print(state,'S')
                # print(destination,'d')
                if destination in ['LA PAZ', 'BENITO JUAREZ','CALERA']:
                    special_case_flag=1
                if special_case_flag==0:
                    destination_index=dest.index(destination)+1
                else:
                    if destination=='LA PAZ':
                        if state=='BCS':
                            destination_index=102
                        else:
                            destination_index=23
                    elif destination=='CALERA':
                        if state=='ZAC':
                            destination_index=502
                        else:
                            destination_index=514
                    else:
                        if state=='QTR':
                            destination_index=119
                        else:
                            destination_index=133
 
                # print(destination_index,'d')
 
                if site=='015':
                    indexUnity=unities.index(unity_type,4,11)+1
                elif site=='009':
                    indexUnity=unities.index(unity_type,12,26)+1
                elif site=='037':
                    indexUnity=unities.index(unity_type,27,34)+1
                elif site=='140':
                    indexUnity=unities.index(unity_type,35,42)+1
                elif site=='130':
                    indexUnity=unities.index(unity_type,43,50)+1
                elif site=='139':
                    indexUnity=unities.index(unity_type,51,58)+1
                elif site=='151':
                    indexUnity=unities.index(unity_type,59,66)+1
                elif site=='187':
                    indexUnity=unities.index(unity_type,67,74)+1
                elif site=='004':
                    indexUnity=unities.index(unity_type,75,90)+1
                elif site=='051':
                    indexUnity=unities.index(unity_type,91,106)+1
                elif site=='100':
                    indexUnity=unities.index(unity_type,107,114)+1
                elif site=='108':
                    indexUnity=unities.index(unity_type,115,122)+1
                elif site=='116':
                    indexUnity=unities.index(unity_type,123,138)+1
                elif site=='065':
                    indexUnity=unities.index(unity_type,139,146)+1
                elif site=='016':
                    indexUnity=unities.index(unity_type,147,155)+1
                elif site=='083':
                    indexUnity=unities.index(unity_type,156,159)+1
                elif site=='186':
                    indexUnity=unities.index(unity_type,160,167)+1
                elif site=='002':
                    indexUnity=unities.index(unity_type,168,176)+1
                elif site=='014':
                    indexUnity=unities.index(unity_type,177,184)+1
                elif site=='019':
                    indexUnity=unities.index(unity_type,185,193)+1
                elif site=='035':
                    indexUnity=unities.index(unity_type,194,202)+1
                elif site=='024':
                    indexUnity=unities.index(unity_type,203,211)+1
                elif site=='146':
                    indexUnity=unities.index(unity_type,212,219)+1
                elif site=='027':
                    indexUnity=unities.index(unity_type,220,227)+1
                elif site=='031':
                    indexUnity=unities.index(unity_type,228,236)+1
                elif site=='132':
                    indexUnity=unities.index(unity_type,237,244)+1
                elif site=='115':
                    indexUnity=unities.index(unity_type,245,252)+1
                elif site=='145':
                    indexUnity=unities.index(unity_type,253,262)+1
                elif site=='182':
                    indexUnity=unities.index(unity_type,263,270)+1
                elif site=='185':
                    indexUnity=unities.index(unity_type,271,275)+1
 
                # print(destination_index,indexUnity,'VectorMatricial')
                fare_CV=ws.cell(destination_index,indexUnity).value
                tarifas.append([fare_CV, f[7],f[6]])
            else:
                tarifas.append([f[-1],f[7],f[6]])
 
        else:
            if spot_flag==1:
                f.pop(0)
                for i in spots:
                    if i[0]==str(f[0]) and i[1]==f[1] and i[2]==f[2] and i[3]==f[3] and i[4]==str(f[4]) and i[9]==str(f[9]):
                        cv=i[6]
                        while len(cv)<8:
                            cv='0'+cv
                        if cv==f[6]:
                            spot_fares.append(f)
                        else:
                            spot_not_valid.append(f[6])
                count=0
                for s in spot_fares:
                    if str(f[6]) not in s:
                        count+=1
                if count==len(spot_fares):
                    spot_not_valid.append(f[6])
            else: 
                spot_not_valid.append(f[7])
 
    fullData.append(tarifas)
    fullData.append(spot_fares)
    fullData.append(spot_not_valid)
    fullData.append(cv_No_Pro)
    # dest = shutil.copy(filename, r'C:\Users\aperalda\Documents')
    print('Tarifas array: ')
    return fullData
 
def validationTarif(filename,mail,msg):
    wb = load_workbook(filename, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    row_count=len(ws['B'])
    for i in range(2,13):
        formato.append(ws.cell(8,i).value.upper())
    # fecha=str(ws.cell(7,5).value).split()[0]
    # if fecha==str(datetime.date.today()):
    #     pass
    if re.match('NO.?',formato[0]) and re.match('.*?ID.*?OTM', formato[1]) and re.match('.*?L[IÍ]NEA', formato[2]) and re.match('.*?PREDIO', formato[3]) and re.match('.*?UNIDAD', formato[4]) and re.match('C[P]?([ÓO]DIGO POSTAL)?', formato[5]) and re.match('POBLACI[ÓO]N', formato[6]) and re.match('C[V]?(ONTROL VEH[ÍI]CULAR)?', formato[7]) and re.match('.*?TARIFA', formato[8]) and re.match('AUTORIZA', formato[9]) and re.match('.*?IMPORTE', formato[10]):
        pass
    else:
        forwardFormatMailError(mail)
    # if type(None) in formato:
    #     pass
    filas=[]
    noProcessedCV=[]
    for i in range(9,row_count+1):
        fila=[]
        for a in range(2,13):
            if ws.cell(i,a).value != None:
                if ws.cell(i,a).value !='':
                    if a==9:
                        cv=str(ws.cell(i,a).value)
                        while len(cv)<8:
                            cv='0'+cv
                        fila.append(cv)
                        continue
                    else:
                        fila.append(ws.cell(i,a).value)
                else:
                    fila.append('')
            else:
                fila.append('')
        if '' in fila:
            noProcessedCV.append(fila[7])
        else:
            filas.append(fila)
 
    print('tarifas a procesar: ', len(filas))
    print('tarifas no procesadas: ', len(noProcessedCV),'\n')
    wb.close()
    tarifas_CV=tariifario(filas,filename,msg)
    tarifas_CV.append(noProcessedCV)
    return tarifas_CV

#-------------------------------------------------------------------------------------------------------------
def validationSPOT(xlsxFile,arr_To_Validate,mail):
 
    inconsistantData=[]
    cv_No_Pro=arr_To_Validate[-1]
    arr_To_Validate=arr_To_Validate[:-1]
    wb = load_workbook(xlsxFile, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    for i in range(2,13):
        formato.append(ws.cell(8,i).value.upper())
    # fecha=str(ws.cell(7,5).value).split()[0]
    # if fecha==str(datetime.date.today()):
    #     pass
    if re.match('NO.?',formato[0]) and re.match('.*?ID.*?OTM', formato[1]) and re.match('.*?L[IÍ]NEA', formato[2]) and re.match('.*?PREDIO', formato[3]) and re.match('.*?UNIDAD', formato[4]) and re.match('C[P]?([ÓO]DIGO POSTAL)?', formato[5]) and re.match('POBLACI[ÓO]N', formato[6]) and re.match('C[V]?(ONTROL VEH[ÍI]CULAR)?', formato[7]) and re.match('.*?TARIFA', formato[8]) and re.match('AUTORIZA', formato[9]) and re.match('.*?IMPORTE', formato[10]):
        pass
    else:
        forwardFormatMailError(mail)
    # if type(None) in formato:
    #     pass
 
    inconsistants=[]
 
    for i in range(len(arr_To_Validate)):
        col=3
        inconsistency=0
        row=arr_To_Validate[i][-1]
        arr_To_Validate[i].pop(-1)
        for s in range(len(arr_To_Validate[0])):
            if s==8:
                fareXlsx=ws.cell(row,col).value
                fareXlsx.replace(' ','')
                if str(fareXlsx)!=arr_To_Validate[i][s]:
                    inconsistency=1
                    print(arr_To_Validate[i][s], ws.cell(row,col).value,'inconsistencia \n')
            elif str(ws.cell(row,col).value)!=str(arr_To_Validate[i][s]):
                inconsistency=1
                print(arr_To_Validate[i][s], ws.cell(row,col).value,'inconsistencia\n')
            col+=1
        if inconsistency==0:
            pass
        else:
            inconsistantData.append(arr_To_Validate[i][6])
            inconsistants.append(arr_To_Validate[i])
 
    for r in inconsistants:
        arr_To_Validate.remove(r)
    
    for element in arr_To_Validate:
        cv=str(element[6])
        while len(cv)<8:
            cv='0'+cv
        element[6]=cv
    
    for ele in inconsistantData:
        cv=str(ele)
        while len(cv)<8:
            cv='0'+cv
        ele=cv
 
    for el in cv_No_Pro:
        cv=str(el)
        while len(cv)<8:
            cv='0'+cv
        el=cv
    if len(arr_To_Validate)==0:
        arr_To_Validate.append([])
    
    spots_processed=[]
    spots_processed.append(arr_To_Validate)
    spots_processed.append(inconsistantData)
    spots_processed.append(cv_No_Pro)
    wb.close()
    print('Spot array:')
    return spots_processed
                
    
 
def forwardFormatMailError(mail):
    reply=mail.Forward()
    sender=mail.Sender
    newBody = "Formato Incorrecto"
    reply.HTMLBody = newBody + reply.HTMLBody
    reply.To=sender
    reply.Send()
    print('Mandado error de formato en xlsx')
 
# def forwardDateError(mail):
#     reply=mail.Forward()
#     sender=mail.Sender
#     print(sender)
#     newBody='Archivo expirado'
#     reply.HTMLBody = newBody + reply.HTMLBody
#     reply.To=sender
#     reply.Send()
#     print('Mandado')
 
def createReply(email):
    reply=email.Forward()
    sender=email.Sender
    print(sender)
    newBody='El número de archivos adjuntos no es el esperado'
    reply.HTMLBody = newBody + reply.HTMLBody
    reply.To=sender
    reply.Send()
    print('Mandado error de archivo adjunto')
 
def retrieval():
    flag=0
 
    global flagTime
    global nowT
    print('actualizando \n')
 
    if flagTime==1:
        timeElapsed=datetime.datetime.now()-nowT
        timeElapsed=timeElapsed.seconds
        timeToWait=60-timeElapsed
        print('Actualización de bandeja en: ',timeToWait, ' segundos')
        time.sleep(timeToWait)
        flagTime=0
 
    nowT=datetime.datetime.now()
    inbox = outlook.GetDefaultFolder(6)                         # "6" refers to the index of a folder - in this case the inbox.                                      
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)
 
    if nowT+datetime.timedelta(seconds=60)>datetime.datetime.now() and flagTime==0:
        flagTime=0
        print('leyendo\n')
        for message in messages:
            if message.Unread==True and 'TARIFAS RPA' in message.Subject.upper():
                numberOfAttachments=len(message.Attachments)
                if numberOfAttachments<=2:
                    print('Guardando\n')
                    flag=1
                    msg=''
                    if numberOfAttachments==2:
                        file_name1=r'D:\Descargas' + '\\'+ message.Attachments[0].FileName
                        file_name2=r'D:\Descargas' + '\\'+ message.Attachments[1].FileName
                        message.Attachments[0].SaveAsFile(file_name1)
                        message.Attachments[1].SaveAsFile(file_name2)
                        docs=[file_name1,file_name2]
                        for i in docs:
                            if '.msg' in i:
                                msg=i
                            else:
                                file_name1=i
                    else:
                        file_name1=r'D:\Descargas' + '\\'+ message.Attachments[0].FileName
                        message.Attachments[0].SaveAsFile(file_name1)
                        
                    message.Unread = False
                    print('Mensaje leído\n')
                    break
                else:
                    message.Unread=False
                    time.sleep(1)
                    print('Error en attachment\n')
                    createReply(message)
            
 
            elif message.Unread==True and 'SPOT RPA' in message.Subject.upper():
                numberOfAttachments=len(message.Attachments)
                if numberOfAttachments==2:
                    print('Guardando\n')
                    flag=2
                    file_name1=r'D:\Descargas' + '\\'+message.Attachments[0].FileName
                    file_name2=r'D:\Descargas' + '\\'+message.Attachments[1].FileName
                    message.Attachments[0].SaveAsFile(file_name1)
                    message.Attachments[1].SaveAsFile(file_name2)
                    docs=[file_name1,file_name2]
                    for i in docs:
                        if '.msg' in i:
                            msg=i
                        else:
                            xlsxFile=i
                    message.Unread = False
                    print('Mensaje leído\n')
                    break
                else:
                    message.Unread=False
                    print('Error en attachment\n')
                    createReply(message,numberOfAttachments)
            
 
    if flag==1:
        flag=0
        masterArray=validationTarif(file_name1, message,msg)
        os.remove(file_name1)
        return masterArray
    elif flag==2:
        flag=0
        masterArray=validationSPOT(xlsxFile,outlookItem(msg),message)
        os.remove(xlsxFile)
        os.remove(msg)
        return masterArray
 
    else:
        print('No se encontró alta de tarifas\n')
 
while 1:
    print(retrieval())
    flagTime=1
