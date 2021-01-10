import win32com.client as win32
import datetime,re,os,shutil,time
from openpyxl import load_workbook
from outlook_msg import Message
 
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
account= win32.Dispatch("Outlook.Application").Session.Accounts
flagTime=0
formato=[]
tfares=r'C:\Users\aperalda\Documents\AltaDeTarifas\TarifMaster.xlsx'

def txt_to_str(route):
    f = open(route, mode="r", encoding="utf-8")
    content = f.read()
    f.close()
    return str(content)

def createReply(email,mailBody,flag,acceptedCV):
    reply=email.Forward()
    sender='alejandro.peraltad@dhl.com'#+'; alta.tarifas.sourcing@dhl.com'
    if flag==0:
        reply.HTMLBody = mailBody +'\n' +reply.HTMLBody
    else:
        init='<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"; content-align:left> <img src="llpc.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV procesados</h2><table style="width:33%"><tr><th>CV Aceptados</th></tr>'
        for b in acceptedCV:
            init=init+'<tr><td>%s</td></tr>' %(b)
        init=init+'</table></body></html>'
        reply.HTMLBody = init + reply.HTMLBody
    images_path = "C:\\Users\\aperalda\\Documents\\RAudit\\img\\"
    reply.Attachments.Add(Source= images_path+"llpc.PNG")
    reply.To=sender
    reply.Send()

def noProcessed(noProcessedCV,subject):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    destinatarios='alejandro.peraltad@dhl.com'#'alta.tarifas.sourcing@dhl.com'
    mail.To = destinatarios
    mail.Subject=subject
    f = open(r'C:\Users\aperalda\Documents\Tarifas\mail.txt', 'w', encoding='UTF-8')
    f.write('<!DOCTYPE html> <html> <head> <title>FORMATO TEXT</title> </head> <style> table, th, td {border: 1px solid black;border-collapse: collapse; text-align: center;}</style><meta charset="UTF-8"> <body style="background-color:#FFE406"> <img src="llpc.png" alt="DHL LOGO.png"> <h2 style="font-family:verdana;text-align:center;">CV no procesados</h2> <p style="font-family:verdana;">Esta es una alerta informativa sobre CV no capturados correctamente:</p><table style="width:100%">')
    if len(noProcessedCV)==3:
        if len(noProcessedCV[1])!=0 or len(noProcessedCV[2])!=0:
            f.write('<tr><th>Inconsistencias</th><th>Falta de datos en correo adjunto</th></tr>')
            if len(noProcessedCV[1])>len(noProcessedCV[2]):
                len_larger=len(noProcessedCV[1])
                larger=noProcessedCV[1]
            else:
                len_larger=len(noProcessedCV[2])
                larger=noProcessedCV[2]
            for i in range(len_larger):
                try:
                    f.write("<tr> <td>%s</td><td>%s</td></tr>\n" %(noProcessedCV[1][i][6], noProcessedCV[2][i]))
                except:
                    if larger==noProcessedCV[1]:
                        f.write("<tr> <td>%s</td><td>%s</td></tr>\n" %(noProcessedCV[1][i][6],''))
                    else:
                        f.write("<tr> <td>%s</td><td>%s</td></tr>\n" %('', noProcessedCV[2][i]))
    elif len(noProcessedCV)==5:
        f.write('<tr><th>Inconsistencias</th><th>Falta de datos en correo adjunto</th><th>Falta de datos en csv</th></tr>')
        if len(noProcessedCV[2])>=len(noProcessedCV[3]) and len(noProcessedCV[2])>=len(noProcessedCV[4]):
            larger=len(noProcessedCV[2])
            while len(noProcessedCV[2])>len(noProcessedCV[3]):
                noProcessedCV[3].append('')
            while len(noProcessedCV[2])>len(noProcessedCV[4]):
                noProcessedCV[4].append('')
        elif len(noProcessedCV[3])>=len(noProcessedCV[2]) and len(noProcessedCV[3])>=len(noProcessedCV[4]):
            larger=len(noProcessedCV[3])
            while len(noProcessedCV[3])>len(noProcessedCV[2]):
                noProcessedCV[2].append('')
            while len(noProcessedCV[3])>len(noProcessedCV[4]):
                noProcessedCV[4].append('')
        else:
            larger=len(noProcessedCV[4])
            while len(noProcessedCV[4])>len(noProcessedCV[2]):
                noProcessedCV[2].append('')
            while len(noProcessedCV[4])>len(noProcessedCV[3]):
                noProcessedCV[3].append('')
        for i in range(larger):
            f.write("<tr><td>%s</td><td>%s</td><td>%s</td></tr>" %(noProcessedCV[2][i],noProcessedCV[3][i],noProcessedCV[4][i]))
    f.write('</table></body></html>')
    f.close()
    body=txt_to_str(r'C:\Users\aperalda\Documents\Tarifas\mail.txt')
    mail.HTMLBody = body
    images_path = "C:\\Users\\aperalda\\Documents\\RAudit\\img\\"
    mail.Attachments.Add(Source= images_path+"llpc.PNG")
    mail.Send()
    print('No procesados enviado\n')

def outlookItem(fname):
    cv_SPOT_No_Processed=[]
    newCont=''
    line=''
    flag=0
    with open(fname) as msg_file:
        msg = Message(msg_file)
    content=msg.body

    for c in content:
        if c!='\n' and c!='\r':
            line+=c
        else:
            flag=1
            pass
        if flag==1:
            if c!='\n' and c!='\r':
                lastChar=line[-1]
                line=line[:-1]
                line=line+'\n'
                line=line+lastChar
                flag=0
            else:
                pass
    line=line.split('\n')
    line = list(map(lambda x:x.upper(),line))
    for ele in range(len(line)):
        if '\t' in line[ele]:
            line.insert(ele,' ')
            line[ele+1]=line[ele+1].replace('\t','')
    data=[]
    fare=[]
    comData=[]
    fullData=[]
    for i in range(len(line)):
        if '$' in line[i]:
            fare.append(i)
    lastFareIndex=fare[-1]
    if 'NÚM SP ID OTM' in line:
        start=line.index('NÚM SP ID OTM')+10
    for i in range(start,lastFareIndex+1,10):
        singleData=[line[i],line[i+1],line[i+2],line[i+3],line[i+4],line[i+5],line[i+6],line[i+7],line[i+8],line[i+9]]
        if ' ' in singleData:
            cv_SPOT_No_Processed.append(line[i+6])
        else:
            comData.append(singleData)
    for i in comData:
        fareInd=i[-1].replace(' ','')
        fareInd=fareInd.replace('$','')
        fareInd=fareInd.replace(',','')
        fareInd=fareInd[0:-3]
        i.pop(-1)
        i.insert(len(i),fareInd)
    
    fullData.append(comData)
    comData.append(cv_SPOT_No_Processed)
    return comData
  
def tariff(filas,filename,msg,mail):
    
    cvProcessed=[]
    trf=load_workbook(tfares)
    ws = trf.worksheets[0]
    origin_Nomenclature_row = list(ws.rows)[1]
    destination_site_col=list(ws.columns)[2]
    unity_row=list(ws.rows)[2]
    dest=[cell.value for cell in destination_site_col]
    unities=[cell.value for cell in unity_row]
    tarifas=[]
    spot_fares=[]
    spot_not_valid=[]
    cv_No_Pro=[]
    fullData=[]
    count=0
 
    try:
        spots=outlookItem(msg)
        spot_flag=1
        cv_No_Pro=spots[-1]
        for s in range(len(cv_No_Pro)):
            cv_s=cv_No_Pro[s]
            while len(cv_s)<8:
                cv_s='0'+cv_s
            cv_No_Pro[s]=cv_s
        spots=spots[:-1]
    except:
        spot_flag=0
        print('No hay elementos para validar SPOT\n')
                
    for f in filas:
        if f[7] == 'TARIFA NORMAL':
            if not re.match('050|128|148', f[3][0:3]):
                special_case_flag=0
                site=f[2].split()[0]
                state=f[5].split('/')[0].strip()
                destination=f[5].split('/')[1].strip()
                unity_type=f[3]
                # print(unity_type,'Unity type') print(site,'o') print(state,'S') print(destination,'d')
                if destination in ['LA PAZ', 'BENITO JUAREZ','CALERA']:
                    special_case_flag=1
                if special_case_flag==0:
                    destination_index=dest.index(destination)+1
                else:
                    if destination=='LA PAZ':
                        if state=='BCS':
                            destination_index=102
                        else:
                            destination_index=23
                    elif destination=='CALERA':
                        if state=='ZAC':
                            destination_index=502
                        else:
                            destination_index=514
                    else:
                        if state=='QTR':
                            destination_index=119
                        else:
                            destination_index=133
                # print(destination_index,'d')

                if site=='015':
                    indexUnity=unities.index(unity_type,4,11)+1
                elif site=='009':
                    indexUnity=unities.index(unity_type,12,26)+1
                elif site=='037':
                    indexUnity=unities.index(unity_type,27,34)+1
                elif site=='140':
                    indexUnity=unities.index(unity_type,35,42)+1
                elif site=='130':
                    indexUnity=unities.index(unity_type,43,50)+1
                elif site=='139':
                    indexUnity=unities.index(unity_type,51,58)+1
                elif site=='151':
                    indexUnity=unities.index(unity_type,59,66)+1
                elif site=='187':
                    indexUnity=unities.index(unity_type,67,74)+1
                elif site=='004':
                    indexUnity=unities.index(unity_type,75,90)+1
                elif site=='051':
                    indexUnity=unities.index(unity_type,91,106)+1
                elif site=='100':
                    indexUnity=unities.index(unity_type,107,114)+1
                elif site=='108':
                    indexUnity=unities.index(unity_type,115,122)+1
                elif site=='116':
                    indexUnity=unities.index(unity_type,123,138)+1
                elif site=='065':
                    indexUnity=unities.index(unity_type,139,146)+1
                elif site=='016':
                    indexUnity=unities.index(unity_type,147,155)+1
                elif site=='083':
                    indexUnity=unities.index(unity_type,156,159)+1
                elif site=='186':
                    indexUnity=unities.index(unity_type,160,167)+1
                elif site=='002':
                    indexUnity=unities.index(unity_type,168,176)+1
                elif site=='014':
                    indexUnity=unities.index(unity_type,177,184)+1
                elif site=='019':
                    indexUnity=unities.index(unity_type,185,193)+1
                elif site=='035':
                    indexUnity=unities.index(unity_type,194,202)+1
                elif site=='024':
                    indexUnity=unities.index(unity_type,203,211)+1
                elif site=='146':
                    indexUnity=unities.index(unity_type,212,219)+1
                elif site=='027':
                    indexUnity=unities.index(unity_type,220,227)+1
                elif site=='031':
                    indexUnity=unities.index(unity_type,228,236)+1
                elif site=='132':
                    indexUnity=unities.index(unity_type,237,244)+1
                elif site=='115':
                    indexUnity=unities.index(unity_type,245,252)+1
                elif site=='145':
                    indexUnity=unities.index(unity_type,253,262)+1
                elif site=='182':
                    indexUnity=unities.index(unity_type,263,270)+1
                elif site=='185':
                    indexUnity=unities.index(unity_type,271,275)+1
 
                # print(destination_index,indexUnity,'VectorMatricial')
                fare_CV=ws.cell(destination_index,indexUnity).value
                f.pop(-1)
                f.insert(len(f),fare_CV)
                tarifas.append(f)
            else:
                tarifas.append(f)
 
        else:
            if spot_flag==1:
                for i in spots:
                    if i[0]==str(f[0]) and i[1]==f[1] and i[2]==f[2] and i[3]==f[3] and i[4]==str(f[4]) and i[9]==str(f[9]):
                        cv=i[6]
                        while len(cv)<8:
                            cv='0'+cv
                        if cv==f[6]:
                            spot_fares.append(f)
                        else:
                            if str(f[6]) not in cv_No_Pro:
                                spot_not_valid.append(f[6])
                count=0
                for s in spot_fares:
                    if str(f[6]) not in s:
                        count+=1
                if count==len(spot_fares) and f[6] not in cv_No_Pro:
                    spot_not_valid.append(f[6])
            else:
                if str(f[6]) not in cv_No_Pro: 
                    spot_not_valid.append(f[6])
    
    for g in tarifas:
        cvProcessed.append(g[6])
    for h in spot_fares:
        cvProcessed.append(h[6])

    fullData.append(tarifas)
    fullData.append(spot_fares)
    fullData.append(spot_not_valid)
    fullData.append(cv_No_Pro)
    createReply(mail,'Tarifas aceptadas: ',1,cvProcessed)
    return fullData
 
def validationTarif(filename,mail,msg):
    wb = load_workbook(filename, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    row_count=len(ws['B'])
    for i in range(2,13):
        formato.append(ws.cell(8,i).value.upper())

    if re.match('NO.?',formato[0]) and re.match('.*?ID.*?OTM', formato[1]) and re.match('.*?L[IÍ]NEA', formato[2]) and re.match('.*?PREDIO', formato[3]) and re.match('.*?UNIDAD', formato[4]) and re.match('C[P]?([ÓO]DIGO POSTAL)?', formato[5]) and re.match('POBLACI[ÓO]N', formato[6]) and re.match('C[V]?(ONTROL VEH[ÍI]CULAR)?', formato[7]) and re.match('.*?TARIFA', formato[8]) and re.match('AUTORIZA', formato[9]) and re.match('.*?IMPORTE', formato[10]):
        pass
    else:
        createReply(mail,"Formato Incorrecto",0,'')

    filas=[]
    noProcessedCV=[]
    for i in range(9,row_count+1):
        fila=[]
        for a in range(3,13):
            if ws.cell(i,a).value != None:
                if ws.cell(i,a).value !='':
                    if a==9:
                        cv=str(ws.cell(i,a).value)
                        while len(cv)<8:
                            cv='0'+cv
                        fila.append(cv)
                        continue
                    else:
                        fila.append(ws.cell(i,a).value)
                else:
                    fila.append('')
            else:
                fila.append('')
        if '' in fila:
            noProcessedCV.append(fila[6])
        else:
            filas.append(fila)

    wb.close()
    tarifs_CV=tariff(filas,filename,msg,mail)
    tarifs_CV.append(noProcessedCV)
    return tarifs_CV

def validationSPOT(xlsxFile,arr_To_Validate,mail):
 
    inconsistantData=[]
    cv_No_Pro=arr_To_Validate[-1]
    arr_To_Validate=arr_To_Validate[:-1]
    wb = load_workbook(xlsxFile, data_only=True)       #Determina el numero de filas
    ws = wb.worksheets[1]
    row_count=len(ws['B'])

    for i in range(2,13):
        formato.append(ws.cell(8,i).value.upper())

    if re.match('NO.?',formato[0]) and re.match('.*?ID.*?OTM', formato[1]) and re.match('.*?L[IÍ]NEA', formato[2]) and re.match('.*?PREDIO', formato[3]) and re.match('.*?UNIDAD', formato[4]) and re.match('C[P]?([ÓO]DIGO POSTAL)?', formato[5]) and re.match('POBLACI[ÓO]N', formato[6]) and re.match('C[V]?(ONTROL VEH[ÍI]CULAR)?', formato[7]) and re.match('.*?TARIFA', formato[8]) and re.match('AUTORIZA', formato[9]) and re.match('.*?IMPORTE', formato[10]):
        pass
    else:
        createReply(mail,"Formato Incorrecto",0,'')
 
    consistants=[]
    for i in range(9,row_count+1):
        if 'SPOT' in str(ws.cell(i,10).value):
            for a in arr_To_Validate:
                if a[0]==str(ws.cell(i,3).value) and a[1]==ws.cell(i,4).value and a[2]==ws.cell(i,5).value and a[3]==ws.cell(i,6).value and a[4]==str(ws.cell(i,7).value) and a[6]==str(ws.cell(i,9).value) and a[-1]==str(ws.cell(i,12).value):
                    consistants.append(a)
                else:
                    pass

    for r in arr_To_Validate:
        if r not in consistants and r[6] not in cv_No_Pro:
            inconsistantData.append(r)

    for element in consistants:
        cv=str(element[6])
        while len(cv)<8:
            cv='0'+cv
        element[6]=cv
    
    for ele in inconsistantData:
        cv=str(ele[6])
        while len(cv)<8:
            cv='0'+cv
        ele[6]=cv
 
    for el in range(len(cv_No_Pro)):
        cv=str(cv_No_Pro[el])
        while len(cv)<8:
            cv='0'+cv
        cv_No_Pro[el]=cv

    if len(arr_To_Validate)==0:
        arr_To_Validate.append([])
    
    spots_processed=[]
    spots_accepted=[]
    for y in consistants:
        spots_accepted.append(y[6])
    spots_processed.append(consistants)
    spots_processed.append(inconsistantData)
    spots_processed.append(cv_No_Pro)
    wb.close()
    createReply(mail,'Spots aceptados: ',1,spots_accepted)
    return spots_processed
                
 
def retrieval():
    flag=0
    inbox = outlook.GetDefaultFolder(6)                         # "6" refers to the index of a folder - in this case the inbox.                                      
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)
    print('leyendo bandeja \n')
    for message in messages:
        if message.Unread==True and 'TARIFAS RPA' in message.Subject.upper():
            numberOfAttachments=len(message.Attachments)
            if numberOfAttachments<=2:
                print('Guardando\n')
                flag=1
                msg=''
                if numberOfAttachments==2:
                    file_name1=r'C:\Users\aperalda\Downloads' + '\\'+ message.Attachments[0].FileName
                    file_name2=r'C:\Users\aperalda\Downloads' + '\\'+ message.Attachments[1].FileName
                    message.Attachments[0].SaveAsFile(file_name1)
                    message.Attachments[1].SaveAsFile(file_name2)
                    docs=[file_name1,file_name2]
                    for i in docs:
                        if '.msg' in i:
                            msg=i
                        else:
                            file_name1=i
                else:
                    file_name1=r'C:\Users\aperalda\Downloads' + '\\'+ message.Attachments[0].FileName
                    message.Attachments[0].SaveAsFile(file_name1)

                message.Unread = False
                print('Mensaje leído\n')
                break
            else:
                message.Unread=False
                time.sleep(1)
                print('Error en attachment\n')
                createReply(message,'El número de archivos adjuntos no es el esperado',0,'')
                break

        elif message.Unread==True and 'SPOT RPA' in message.Subject.upper():
            numberOfAttachments=len(message.Attachments)
            if numberOfAttachments==2:
                print('Guardando\n')
                flag=2
                file_name1=r'C:\Users\aperalda\Downloads' + '\\'+message.Attachments[0].FileName
                file_name2=r'C:\Users\aperalda\Downloads' + '\\'+message.Attachments[1].FileName
                message.Attachments[0].SaveAsFile(file_name1)
                message.Attachments[1].SaveAsFile(file_name2)
                docs=[file_name1,file_name2]
                for i in docs:
                    if '.msg' in i:
                        msg=i
                    else:
                        xlsxFile=i
                message.Unread = False
                print('Mensaje leído\n')
                break
            else:
                message.Unread=False
                print('Error en attachment\n')
                createReply(message,'El número de archivos adjuntos no es el esperado',0,'')
                break
            
 
    if flag==1:
        flag=0
        final_fares_Array=validationTarif(file_name1, message,msg)
        noProcessed(final_fares_Array,'CV no procesados')
        os.remove(file_name1)
        print('Tarifas array: ')
        print(final_fares_Array)
        return final_fares_Array
    elif flag==2:
        flag=0
        final_fares_Array=validationSPOT(xlsxFile,outlookItem(msg),message)
        noProcessed(final_fares_Array,'CV no procesados')
        os.remove(xlsxFile)
        os.remove(msg)
        print('Spot array:')
        print(final_fares_Array)
        return final_fares_Array
 
    else:
        print('No se encontró alta de tarifas\n')
 
while 1:
    final_fares=retrieval()
    print('Actualizando')
    time.sleep(60)
